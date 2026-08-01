from __future__ import annotations

import io
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .analysis import analysis_to_dataframe


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[get_column_letter(column)].width = min(max_length + 2, 48)


def _style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _result_dataframe(analysis: dict[str, Any]) -> pd.DataFrame:
    result_csv = (analysis.get("result_csv") or "").strip()
    if not result_csv:
        return pd.DataFrame()
    try:
        return pd.read_csv(io.StringIO(result_csv))
    except Exception:
        return analysis_to_dataframe(analysis)


def build_workbook(analysis: dict[str, Any], include_details: bool = True) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    df = _result_dataframe(analysis)
    if df.empty:
        ws["A1"] = "No preview rows available"
    else:
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        _style_header(ws, 1)
        ws.freeze_panes = "A2"
        _autosize(ws)

    summary = wb.create_sheet("Analysis")
    summary.append(["Metric", "Value"])
    summary.append(["Rows", analysis["summary"]["rows"]])
    summary.append(["Columns", analysis["summary"]["columns"]])
    summary.append(["Missing Cells", analysis["summary"]["missing_cells"]])
    summary.append(["Duplicate Rows", analysis["summary"]["duplicate_rows"]])
    summary.append(["Memory (bytes)", analysis["summary"]["memory_bytes"]])
    summary.append(["SQL Query Kind", analysis["sql"]["query_kind"]])
    summary.append(["SQL Statements", analysis["sql"]["statement_count"]])
    summary.append(["SQL Token Count", analysis["sql"]["token_count"]])
    summary.append(["VBA Procedures", analysis["vba"]["procedure_count"]])
    summary.append(["VBA Complexity Score", analysis["vba"]["complexity_score"]])
    _style_header(summary, 1)
    _autosize(summary)

    if include_details:
        cols = wb.create_sheet("Columns")
        cols.append(["Column", "Dtype", "Non-null", "Missing", "Unique", "Null Ratio"])
        for col in analysis.get("columns", []):
            cols.append([
                col["name"],
                col["dtype"],
                col["non_null"],
                col["missing"],
                col["unique"],
                col["null_ratio"],
            ])
        _style_header(cols, 1)
        _autosize(cols)

        vba = wb.create_sheet("VBA")
        vba.append(["Metric", "Value"])
        for key in ("line_count", "non_empty_lines", "blank_lines", "character_count", "procedure_count", "attribute_count", "comment_lines", "on_error_count", "complexity_score"):
            vba.append([key, analysis["vba"].get(key)])
        _style_header(vba, 1)
        _autosize(vba)

    return wb


def workbook_to_bytes(wb: openpyxl.Workbook, suffix: str) -> bytes:
    with NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(tmp_path)
        return tmp_path.read_bytes()
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass


def analysis_to_csv_bytes(analysis: dict[str, Any]) -> bytes:
    df = _result_dataframe(analysis)
    return df.to_csv(index=False).encode("utf-8-sig")


def analysis_to_xlsx_bytes(analysis: dict[str, Any]) -> bytes:
    wb = build_workbook(analysis, include_details=True)
    return workbook_to_bytes(wb, ".xlsx")


def analysis_to_xlsm_bytes(analysis: dict[str, Any]) -> bytes:
    wb = build_workbook(analysis, include_details=True)
    # NOTE: This writes a macro-enabled workbook container. If you need embedded VBA modules,
    # extend this function to inject a vbaProject.bin from a trusted template workbook.
    return workbook_to_bytes(wb, ".xlsm")
